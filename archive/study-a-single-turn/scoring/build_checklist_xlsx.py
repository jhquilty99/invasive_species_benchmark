"""Generate scoring/checklist.xlsx from scoring/checklist.jsonl for expert/reviewer use.
Regenerate by rerunning this script whenever checklist.jsonl changes; it always rebuilds
from scratch, so any in-progress annotations in the existing .xlsx are not preserved
across a rerun.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
CHECKLIST_PATH = REPO_ROOT / "scoring" / "checklist.jsonl"
OUT_PATH = REPO_ROOT / "scoring" / "checklist.xlsx"

SUMMARY_COLUMNS = ["item_id", "num_claims", "num_critical", "num_standard", "num_triggers", "max_harm_level"]
CLAIMS_COLUMNS = ["item_id", "claim_id", "weight", "text"]
TRIGGERS_COLUMNS = ["item_id", "trigger_id", "harm_level", "condition"]

HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
HEADER_FONT = Font(bold=True)


def load_items() -> list[dict]:
    with CHECKLIST_PATH.open(encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def max_harm_level(triggers: list[dict]) -> str | None:
    if not triggers:
        return None
    return max((t["harm_level"] for t in triggers), key=lambda h: int(h[1:]))


def write_sheet(ws, columns: list[str], rows: list[tuple], widths: dict[str, int]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)
    ws.freeze_panes = "A2"
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if "text" in columns:
            ws.cell(row=row_idx, column=columns.index("text") + 1).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        if "condition" in columns:
            ws.cell(row=row_idx, column=columns.index("condition") + 1).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def build() -> None:
    items = load_items()

    summary_rows = []
    claim_rows = []
    trigger_rows = []
    for item in items:
        item_id = item["item_id"]
        claims = item.get("claims", [])
        triggers = item.get("harm_triggers", [])
        num_critical = sum(1 for c in claims if c["weight"] == "critical")
        num_standard = sum(1 for c in claims if c["weight"] == "standard")
        summary_rows.append(
            (item_id, len(claims), num_critical, num_standard, len(triggers), max_harm_level(triggers))
        )
        for c in claims:
            claim_rows.append((item_id, c["claim_id"], c["weight"], c["text"]))
        for t in triggers:
            trigger_rows.append((item_id, t["trigger_id"], t["harm_level"], t["condition"]))

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_sheet(
        ws_summary,
        SUMMARY_COLUMNS,
        summary_rows,
        {"item_id": 18, "num_claims": 10, "num_critical": 10, "num_standard": 11, "num_triggers": 11, "max_harm_level": 14},
    )

    ws_claims = wb.create_sheet("Claims")
    write_sheet(ws_claims, CLAIMS_COLUMNS, claim_rows, {"item_id": 18, "claim_id": 22, "weight": 10, "text": 90})

    ws_triggers = wb.create_sheet("Harm Triggers")
    write_sheet(
        ws_triggers, TRIGGERS_COLUMNS, trigger_rows, {"item_id": 18, "trigger_id": 22, "harm_level": 10, "condition": 90}
    )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {len(items)} items ({len(claim_rows)} claims, {len(trigger_rows)} triggers) to {OUT_PATH}")


if __name__ == "__main__":
    build()
