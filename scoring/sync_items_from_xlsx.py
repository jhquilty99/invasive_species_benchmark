"""One-off sync: apply reviewer edits from a downloaded items-review.xlsx (columns A-L only;
later unlabeled columns are ignored as paste spillover) back into data/items.jsonl.

Only overwrites the fields that live in columns A-L: species, category, is_abstention,
abstention_reason, query_text, ground_truth_answer, ground_truth_citation.{source,url,
publication_date}, jurisdiction, jurisdiction_range_flag. item_id, condition_2_documents, and
notes are preserved from the existing items.jsonl since they aren't represented in the sheet.
Items present in items.jsonl but absent from the sheet are left untouched.
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
ITEMS_PATH = REPO_ROOT / "data" / "items.jsonl"

COLUMNS = [
    "item_id",
    "species",
    "category",
    "is_abstention",
    "abstention_reason",
    "query_text",
    "ground_truth_answer",
    "ground_truth_citation_source",
    "ground_truth_citation_url",
    "ground_truth_citation_publication_date",
    "jurisdiction",
    "jurisdiction_range_flag",
]


def load_sheet_rows(xlsx_path: Path) -> dict[str, dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]][: len(COLUMNS)]
    if header != COLUMNS:
        raise SystemExit(f"Unexpected header in {xlsx_path}: {header}")
    rows = {}
    for row in ws.iter_rows(min_row=2, max_col=len(COLUMNS), values_only=True):
        record = dict(zip(COLUMNS, row))
        item_id = record["item_id"]
        if not item_id:
            continue
        rows[item_id] = record
    return rows


def apply_edits(item: dict, edit: dict) -> dict:
    updated = dict(item)
    updated["species"] = edit["species"]
    updated["category"] = edit["category"]
    updated["is_abstention"] = bool(edit["is_abstention"])
    updated["abstention_reason"] = edit["abstention_reason"]
    updated["query_text"] = edit["query_text"]
    updated["ground_truth_answer"] = edit["ground_truth_answer"]
    src = edit["ground_truth_citation_source"]
    url = edit["ground_truth_citation_url"]
    pubdate = edit["ground_truth_citation_publication_date"]
    updated["ground_truth_citation"] = (
        {"source": src, "url": url, "publication_date": pubdate} if (src or url) else None
    )
    updated["jurisdiction"] = edit["jurisdiction"]
    updated["jurisdiction_range_flag"] = bool(edit["jurisdiction_range_flag"])
    return updated


def main(xlsx_path: Path) -> None:
    with ITEMS_PATH.open(encoding="utf-8") as f:
        items = [json.loads(line) for line in f if line.strip()]

    edits = load_sheet_rows(xlsx_path)

    updated_ids = []
    unchanged_ids = []
    new_items = []
    for item in items:
        item_id = item["item_id"]
        if item_id in edits:
            new_items.append(apply_edits(item, edits[item_id]))
            updated_ids.append(item_id)
        else:
            new_items.append(item)
            unchanged_ids.append(item_id)

    missing_in_json = sorted(set(edits) - {it["item_id"] for it in items})
    if missing_in_json:
        raise SystemExit(f"xlsx has item_ids not present in items.jsonl: {missing_in_json}")

    with ITEMS_PATH.open("w", encoding="utf-8", newline="\n") as f:
        for item in new_items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    print(f"Updated {len(updated_ids)} items: {updated_ids}")
    print(f"Left {len(unchanged_ids)} items unchanged (not in sheet): {unchanged_ids}")


if __name__ == "__main__":
    xlsx_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else REPO_ROOT / "scoring" / "items-review.xlsx"
    main(xlsx_arg)
