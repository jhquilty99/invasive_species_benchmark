"""Generate scoring/items-review.xlsx from data/items.jsonl for the freeze-gate review
(SCRATCHPAD.md task 1). Regenerate by rerunning this script whenever items.jsonl changes;
it always rebuilds from scratch, so any in-progress "validated" checkmarks in the existing
.xlsx are not preserved across a rerun.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parent.parent
ITEMS_PATH = REPO_ROOT / "data" / "items.jsonl"
OUT_PATH = REPO_ROOT / "scoring" / "items-review.xlsx"

COLUMNS = [
    "item_id",
    "species",
    "category",
    "is_abstention",
    "abstention_reason",
    "query_text",
    "ground_truth_answer",
    "ground_truth_citation_source",
    "ground_truth_citation_url",
    "ground_truth_citation_publication_date",
    "jurisdiction",
    "jurisdiction_range_flag",
    "condition_2_documents",
    "notes",
    "validated",
]

WIDTHS = {
    "item_id": 16,
    "species": 22,
    "category": 20,
    "is_abstention": 12,
    "abstention_reason": 20,
    "query_text": 50,
    "ground_truth_answer": 70,
    "ground_truth_citation_source": 40,
    "ground_truth_citation_url": 30,
    "ground_truth_citation_publication_date": 16,
    "jurisdiction": 12,
    "jurisdiction_range_flag": 12,
    "condition_2_documents": 40,
    "notes": 20,
    "validated": 10,
}


def load_items():
    with ITEMS_PATH.open(encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def flatten(item):
    citation = item.get("ground_truth_citation") or {}
    docs = item.get("condition_2_documents") or []
    return {
        "item_id": item.get("item_id"),
        "species": item.get("species"),
        "category": item.get("category"),
        "is_abstention": item.get("is_abstention"),
        "abstention_reason": item.get("abstention_reason"),
        "query_text": item.get("query_text"),
        "ground_truth_answer": item.get("ground_truth_answer"),
        "ground_truth_citation_source": citation.get("source"),
        "ground_truth_citation_url": citation.get("url"),
        "ground_truth_citation_publication_date": citation.get("publication_date"),
        "jurisdiction": item.get("jurisdiction"),
        "jurisdiction_range_flag": item.get("jurisdiction_range_flag"),
        "condition_2_documents": "; ".join(docs),
        "notes": item.get("notes"),
        "validated": False,
    }


def build():
    items = load_items()
    wb = Workbook()
    ws = wb.active
    ws.title = "items"

    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS.get(col_name, 18)
    ws.freeze_panes = "A2"

    for row_idx, item in enumerate(items, start=2):
        row = flatten(item)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[col_name])
        for col_name in ("query_text", "ground_truth_answer"):
            ws.cell(row=row_idx, column=COLUMNS.index(col_name) + 1).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    # Boolean checkbox behavior for the validated column: TRUE/FALSE dropdown.
    validated_col_idx = COLUMNS.index("validated") + 1
    validated_col_letter = get_column_letter(validated_col_idx)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{validated_col_letter}2:{validated_col_letter}{len(items) + 1}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(items) + 1}"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {len(items)} rows to {OUT_PATH}")


if __name__ == "__main__":
    build()
